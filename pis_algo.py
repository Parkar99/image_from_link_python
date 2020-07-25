import os
import requests
import re
import urllib
from openpyxl import load_workbook

FILE_NAME = 'image_links.xlsx'

book = load_workbook(FILE_NAME)
sheet = book.active

image_dir = os.path.join(os.path.dirname(__file__), 'images')
if not os.path.exists(image_dir):
    os.mkdir(image_dir)

file_count = 0
for r_number, row in enumerate(sheet.rows):
    sku = sheet.cell(row=r_number + 1, column=1).value
    sku_folder = os.path.join(image_dir, sku)

    row_count = 0
    for cell in row:
        if str(cell.value).startswith('http'):
            if not os.path.exists(sku_folder):
                os.mkdir(sku_folder)

            image = (cell.value).strip()
            image = image[:-1] + '1'

            file_name_data = re.search(
                '((\w|%)+)(\.\w+)+(?!.*(\w+)(\.(\w|%)+)+)', image
            )
            file_name = urllib.parse.unquote(file_name_data.group(0))

            print(f'Found {image} on row {r_number + 1}')
            print(f'Downloading {file_name}')
            image_data = requests.get(image).content
            file_count += 1
            row_count += 1
            with open(f'images/{sku}/{file_name}', 'wb') as file:
                file.write(image_data)

print(f'Total files downloaded = {file_count}')
