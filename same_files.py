from openpyxl import load_workbook

FILE_NAME = 'image_links.xlsx'

book = load_workbook(FILE_NAME)
sheet = book.active

links = []
for r_number, row in enumerate(sheet.rows):
    sku = sheet.cell(row=r_number + 1, column=1).value
    for cell in row:
        if str(cell.value).startswith('http'):
            links.append((sku, str(cell.value)))

for link in links:
    count = list(map(lambda l: l[1], links)).count(link[1])
    if count > 1:
        print(f'{link[1]} found {count} links ({link[0]})')
